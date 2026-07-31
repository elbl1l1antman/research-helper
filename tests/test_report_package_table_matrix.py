from pathlib import Path

import openpyxl

from report_automation_engine.report_package import build_preflight, build_report_package


def make_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "보고서_분석문"
    ws.append(["table_key", "문항/표 제목", "분석문_기본", "최종 사용문", "사용자 수정문", "주요 수치 요약", "검토 상태"])
    ws.append(["T001", "만족도", "만족도는 63.3%로 나타남.", "", "", "63.3%", ""])
    table_ws = wb.create_sheet("보고서_삽입표")
    table_ws.append(["table_key", "title", "category", "weighted_n", "raw_n", "percent", "unit", "source_cell"])
    table_ws.append(["T001", "만족도", "전체", 1200, 1198, 63.25, "%", "D5"])
    table_ws.append(["T001", "만족도", "매우 만족", 353, 351, 29.44, "%", "D6"])
    wb.save(path)


def test_package_tables_include_matrix(tmp_path):
    excel = tmp_path / "sample.xlsx"
    make_workbook(excel)
    package = build_report_package(excel, {"decimal_places": "1"})
    table = package["tables"][0]
    assert table["matrix"][1][1]["display_text"] == "63.3%"
    assert table["matrix"][1][1]["raw_value"] == 63.25
    assert table["roles"]["header"] == 4


def test_preflight_counts_table_matrix_warnings(tmp_path):
    excel = tmp_path / "sample.xlsx"
    make_workbook(excel)
    package = build_report_package(excel, {"decimal_places": "1"})
    preflight = build_preflight(package, [])
    assert "table_matrix_warning_count" in preflight["summary"]
    assert preflight["summary"]["table_matrix_error_count"] == 0


def test_preflight_blocks_section_missing_insert_table_rows(tmp_path):
    excel = tmp_path / "missing_table.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "보고서_분석문"
    ws.append(["table_key", "문항/표 제목", "분석문_기본", "최종 사용문", "사용자 수정문", "주요 수치 요약", "검토 상태"])
    ws.append(["T001", "만족도", "만족도는 63.3%로 나타남.", "", "", "63.3%", ""])
    table_ws = wb.create_sheet("보고서_삽입표")
    table_ws.append(["table_key", "title", "category", "weighted_n", "raw_n", "percent", "unit", "source_cell"])
    table_ws.append(["T999", "만족도", "전체", 1200, 1198, 63.25, "%", "D5"])
    wb.save(excel)

    package = build_report_package(excel, {"decimal_places": "1"})
    preflight = build_preflight(package, [])

    assert any("삽입표 데이터가 없습니다" in item["message"] for item in package["qa"])
    assert preflight["status"] == "blocked"
