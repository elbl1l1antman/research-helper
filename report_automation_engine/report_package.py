"""Build report_package.json and preflight_report.json from Excel outputs."""

from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List

import openpyxl

try:
    from .template_inspector import inspect_template
except ImportError:
    from template_inspector import inspect_template

try:
    from .report_table_matrix import build_table_matrix
except ImportError:
    from report_table_matrix import build_table_matrix


def build_report_package(excel_path: str | Path, meta: Dict[str, Any] | None = None) -> Dict[str, Any]:
    path = Path(excel_path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    qa = read_qa(wb)
    sections = read_sections(wb, qa)
    charts = read_rows(wb, "보고서_차트데이터")
    decimal_places = int(str((meta or {}).get("decimal_places", "1") or "1"))
    tables = group_table_rows(read_rows(wb, "보고서_삽입표"), decimal_places)
    package = {
        "schema_version": "1.0",
        "meta": {
            "source_workbook": str(path),
            "source_file_name": path.name,
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            **(meta or {}),
        },
        "sections": sections,
        "tables": tables,
        "charts": normalize_charts(charts, qa),
        "qa": qa,
    }
    add_contract_qa(package)
    return package


def latest_sheet(wb, prefix: str):
    sheets = [ws for ws in wb.worksheets if ws.title.startswith(prefix)]
    return sheets[-1] if sheets else None


def read_rows(wb, sheet_prefix: str) -> List[Dict[str, Any]]:
    ws = latest_sheet(wb, sheet_prefix)
    if ws is None:
        return []
    headers = [clean(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: List[Dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[i]: values[i] for i in range(min(len(headers), len(values))) if headers[i]}
        if any(clean(value) for value in row.values()):
            rows.append(row)
    return rows


def read_qa(wb) -> List[Dict[str, Any]]:
    return [
        enrich_qa(
            {
                "table_key": clean(row.get("table_key")),
                "type": clean(row.get("qa_type")),
                "severity": clean(row.get("severity")) or "warning",
                "message": clean(row.get("message")),
                "source_range": clean(row.get("source_range")),
                "checked_at": clean(row.get("checked_at")),
            }
        )
        for row in read_rows(wb, "보고서_QA")
    ]


def read_sections(wb, qa: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    sections = []
    seen: set[str] = set()
    for row in read_rows(wb, "보고서_분석문"):
        table_key = clean(row.get("table_key"))
        narrative = first_text(row, "사용자 수정문", "최종 사용문", "분석문_기본")
        flags = sentence_flags(row, narrative)
        if table_key in seen:
            qa.append(issue(table_key, "contract", "error", "table_key가 중복되었습니다."))
        seen.add(table_key)
        sections.append(
            {
                "table_key": table_key,
                "title": clean(row.get("문항/표 제목")),
                "narrative_final": narrative,
                "narrative_source": narrative_source(row),
                "summary": clean(row.get("주요 수치 요약")),
                "review_status": clean(row.get("검토 상태")),
                "qa_flags": flags,
            }
        )
    return sections


def first_text(row: Dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = clean(row.get(key))
        if value:
            return value
    return ""


def narrative_source(row: Dict[str, Any]) -> str:
    for key in ("사용자 수정문", "최종 사용문", "분석문_기본"):
        if clean(row.get(key)):
            return key
    return ""


def sentence_flags(row: Dict[str, Any], narrative: str) -> List[str]:
    flags = []
    if not clean(row.get("문항/표 제목")):
        flags.append("제목 없음")
    if not any(ch.isdigit() for ch in narrative):
        flags.append("수치 없음")
    if narrative and narrative[-1] not in ".다음됨함임남":
        flags.append("종결 표현 확인")
    if len(narrative) < 20:
        flags.append("너무 짧음")
    return flags


def normalize_charts(rows: List[Dict[str, Any]], qa: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    charts = []
    for row in rows:
        value = row.get("value")
        parsed = to_number(value)
        table_key = clean(row.get("table_key"))
        if value not in (None, "") and parsed is None:
            qa.append(issue(table_key, "contract", "error", "차트 value가 숫자가 아닙니다."))
        charts.append(
            {
                "table_key": table_key,
                "series_group": clean(row.get("series_group")),
                "category": clean(row.get("category")),
                "measure": clean(row.get("measure")),
                "value": parsed,
                "display_value": clean(row.get("display_value")),
                "sort_order": to_number(row.get("sort_order")),
                "include_chart": clean(row.get("include_chart")).upper() == "Y",
            }
        )
    return charts


def group_table_rows(rows: List[Dict[str, Any]], decimal_places: int = 1) -> List[Dict[str, Any]]:
    grouped: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        table_key = clean(row.get("table_key"))
        if not table_key:
            continue
        table = grouped.setdefault(table_key, {"table_key": table_key, "title": clean(row.get("title")), "rows": []})
        table["rows"].append(
            {
                "category": clean(row.get("category")),
                "weighted_n": to_number(row.get("weighted_n")),
                "raw_n": to_number(row.get("raw_n")),
                "percent": to_number(row.get("percent")),
                "unit": clean(row.get("unit")),
                "source_cell": clean(row.get("source_cell")),
            }
        )
    return [build_table_matrix(table, decimal_places) for table in grouped.values()]


def add_contract_qa(package: Dict[str, Any]) -> None:
    seen = set()
    for section in package["sections"]:
        key = section["table_key"]
        if not key:
            package["qa"].append(issue("", "contract", "error", "section table_key가 비어 있습니다."))
        if key in seen:
            package["qa"].append(issue(key, "contract", "error", "section table_key가 중복되었습니다."))
        seen.add(key)
        if not section["narrative_final"]:
            package["qa"].append(issue(key, "contract", "error", "최종 문장이 없습니다."))
    if not package["sections"]:
        package["qa"].append(issue("", "contract", "error", "보고서_분석문 산출 시트가 없거나 비어 있습니다."))
    for table in package["tables"]:
        if not table["rows"]:
            package["qa"].append(issue(table["table_key"], "contract", "error", "삽입표 rows가 비어 있습니다."))
        for qa_item in table.get("qa", []):
            package["qa"].append(
                issue(table["table_key"], "table_matrix", qa_item.get("severity", "warning"), qa_item.get("message", ""))
            )


def build_preflight(package: Dict[str, Any], templates: Iterable[tuple[str, str, str]]) -> Dict[str, Any]:
    warnings = [q for q in package["qa"] if clean(q.get("severity")).lower() == "warning"]
    errors = [q for q in package["qa"] if clean(q.get("severity")).lower() == "error"]
    template_reports = []
    for label, path, template_type in templates:
        if path:
            report = inspect_template(path, template_type)
            report["label"] = label
            template_reports.append(report)
            if report["status"] in {"unsupported", "needs_autofix"}:
                errors.append(issue("", "template", "error", f"{label} 템플릿 상태가 {report['status']}입니다."))
            elif report["status"] == "usable_with_warnings":
                warnings.append(issue("", "template", "warning", f"{label} 템플릿 권장 필드가 부족합니다."))

    chart_candidates = [row for row in package["charts"] if row.get("include_chart")]
    table_rows = sum(len(table["rows"]) for table in package["tables"])
    for section in package["sections"]:
        if section["qa_flags"]:
            warnings.append(issue(section["table_key"], "sentence", "warning", ", ".join(section["qa_flags"])))
    if not chart_candidates:
        warnings.append(issue("", "chart", "warning", "차트 후보가 없습니다."))

    status = "blocked" if errors else "ready_with_warnings" if warnings else "ready"
    warning_buckets = count_by(warnings, "review_bucket")
    warning_categories = count_by(warnings, "category")
    table_matrix_warnings = [q for q in warnings if clean(q.get("type")) == "table_matrix"]
    table_matrix_errors = [q for q in errors if clean(q.get("type")) == "table_matrix"]
    return {
        "schema_version": "1.0",
        "status": status,
        "summary": {
            "package_created": True,
            "section_count": len(package["sections"]),
            "chart_candidate_count": len(chart_candidates),
            "table_count": len(package["tables"]),
            "table_row_count": table_rows,
            "qa_warning_count": len(warnings),
            "qa_error_count": len(errors),
            "table_matrix_warning_count": len(table_matrix_warnings),
            "table_matrix_error_count": len(table_matrix_errors),
            "qa_warning_buckets": warning_buckets,
            "qa_warning_categories": warning_categories,
        },
        "template_reports": template_reports,
        "warnings": warnings,
        "errors": errors,
    }


def issue(table_key: str, issue_type: str, severity: str, message: str) -> Dict[str, Any]:
    return enrich_qa({"table_key": table_key, "type": issue_type, "severity": severity, "message": message})


def enrich_qa(item: Dict[str, Any]) -> Dict[str, Any]:
    classification = classify_qa(item.get("type"), item.get("severity"), item.get("message"))
    return {**item, **classification}


def classify_qa(issue_type: Any, severity: Any, message: Any) -> Dict[str, str]:
    issue_type_text = clean(issue_type)
    severity_text = clean(severity).lower()
    message_text = clean(message)

    if severity_text == "error":
        return {
            "category": "blocking_contract",
            "review_action": "fix_required",
            "review_bucket": "improvement_needed",
            "review_bucket_label": "개선 필요",
        }

    if issue_type_text == "표 구조 QA" and "BASE" in message_text.upper():
        return {
            "category": "base_check_needed",
            "review_action": "manual_base_review",
            "review_bucket": "normal_review_warning",
            "review_bucket_label": "정상 검토 경고",
        }

    if issue_type_text == "수치 QA":
        return {
            "category": "no_numeric_points",
            "review_action": "parser_or_manual_review",
            "review_bucket": "improvement_needed",
            "review_bucket_label": "개선 필요",
        }

    if not issue_type_text and ("표 처리" in message_text or message_text.startswith("총 ")):
        return {
            "category": "run_summary",
            "review_action": "reference_only",
            "review_bucket": "info",
            "review_bucket_label": "정보",
        }

    if issue_type_text == "sentence":
        return {
            "category": "sentence_review",
            "review_action": "manual_sentence_review",
            "review_bucket": "normal_review_warning",
            "review_bucket_label": "정상 검토 경고",
        }

    if issue_type_text == "template":
        return {
            "category": "template_warning",
            "review_action": "template_review",
            "review_bucket": "improvement_needed",
            "review_bucket_label": "개선 필요",
        }

    if issue_type_text == "chart":
        return {
            "category": "chart_candidate_warning",
            "review_action": "chart_review",
            "review_bucket": "normal_review_warning",
            "review_bucket_label": "정상 검토 경고",
        }

    return {
        "category": "general_warning",
        "review_action": "manual_review",
        "review_bucket": "normal_review_warning",
        "review_bucket_label": "정상 검토 경고",
    }


def count_by(items: Iterable[Dict[str, Any]], key: str) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    for item in items:
        value = clean(item.get(key)) or "미분류"
        counts[value] = counts.get(value, 0) + 1
    return counts


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def to_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", "").replace("%", "").strip())
    except ValueError:
        return None


def run_cli(argv: List[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Excel 산출 시트를 report package와 preflight로 정규화합니다.")
    parser.add_argument("--excel", required=True)
    parser.add_argument("--package-output", required=True)
    parser.add_argument("--preflight-output", required=True)
    parser.add_argument("--hwp-template", default="")
    parser.add_argument("--ppt-template", default="")
    parser.add_argument("--output-type", default="Excel 산출 시트")
    parser.add_argument("--report-profile", default="")
    parser.add_argument("--style-profile", default="")
    parser.add_argument("--banner", default="")
    parser.add_argument("--decimal-places", default="1")
    args = parser.parse_args(argv)

    package = build_report_package(
        args.excel,
        {
            "output_type": args.output_type,
            "report_profile": args.report_profile,
            "style_profile": args.style_profile,
            "banner": args.banner,
            "decimal_places": args.decimal_places,
        },
    )
    templates = [
        ("HWPX", args.hwp_template, "hwpx_report"),
        ("PPTX", args.ppt_template, "pptx_report"),
    ]
    preflight = build_preflight(package, templates)
    Path(args.package_output).write_text(json.dumps(package, ensure_ascii=False, indent=2), encoding="utf-8")
    Path(args.preflight_output).write_text(json.dumps(preflight, ensure_ascii=False, indent=2), encoding="utf-8")
    print(str(args.package_output))
    print(str(args.preflight_output))
    return 0


if __name__ == "__main__":
    raise SystemExit(run_cli())
