"""Build dashboard_package.json for organization dashboard PPTX output."""

from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List

import openpyxl


def inspect_workbook(excel_path: str | Path, output_path: str | Path) -> Path:
    path = Path(excel_path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(min_row=1, max_row=31, values_only=True))
        if not rows:
            continue
        headers = [clean(value) or f"Column {idx + 1}" for idx, value in enumerate(rows[0])]
        data_rows = rows[1:]
        columns = []
        for idx, header in enumerate(headers):
            values = [row[idx] if idx < len(row) else None for row in data_rows]
            non_empty = [value for value in values if clean(value)]
            numeric = sum(1 for value in non_empty if to_number(value) is not None)
            sample = next((clean(value) for value in non_empty), "")
            columns.append(
                {
                    "name": header,
                    "index": idx + 1,
                    "sample": sample,
                    "inferred_type": "number" if non_empty and numeric >= max(1, len(non_empty) // 2) else "text",
                    "missing_count": max(0, len(data_rows) - len(non_empty)),
                }
            )
        preview = [
            {headers[idx]: clean(value) for idx, value in enumerate(row[: len(headers)])}
            for row in data_rows[:30]
            if any(clean(value) for value in row)
        ]
        sheets.append({"name": ws.title, "columns": columns, "preview": preview})

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        json.dumps({"workbook": str(path), "sheets": sheets}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return output


def build_dashboard_package(
    excel_path: str | Path,
    selection_path: str | Path,
    mapping_path: str | Path,
    page_size: str,
) -> tuple[Dict[str, Any], Dict[str, Any]]:
    source = Path(excel_path)
    selection = load_json(selection_path)
    mapping = load_json(mapping_path)
    page_size = clean(page_size).upper()
    qa: List[Dict[str, Any]] = []

    wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
    sheet_name = clean(selection.get("selected_sheet"))
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
    if ws is None:
        qa.append(issue("selection", "error", "선택한 sheet를 찾을 수 없습니다."))
        rows: List[Dict[str, Any]] = []
        headers: List[str] = []
    else:
        headers, rows = read_sheet_rows(ws, int(selection.get("header_row") or 1))

    selected_columns = [clean(value) for value in selection.get("selected_columns", []) if clean(value)]
    entity_column = clean(selection.get("entity_name_column"))
    selected_names = {clean(value) for value in selection.get("selected_entity_names", []) if clean(value)}
    output_mode = clean(mapping.get("output_mode")) or ("single" if len(selected_names) <= 1 else "batch")

    validate_selection(headers, selected_columns, entity_column, selected_names, page_size, qa)
    validate_mapping(headers, selected_columns, mapping, qa)

    entities = []
    for idx, row in enumerate(rows, start=1):
        entity_name = clean(row.get(entity_column))
        if selected_names and entity_name not in selected_names:
            continue
        entities.append(build_entity(str(idx), entity_name, row, mapping, qa))

    if not entities:
        qa.append(issue("selection", "error", "선택된 기관/기업 행이 없습니다."))

    package = {
        "schema_version": "1.0",
        "meta": {
            "source_workbook": str(source),
            "source_file_name": source.name,
            "selected_sheet": sheet_name,
            "page_size": page_size,
            "output_mode": output_mode,
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        },
        "selection": selection,
        "mapping": mapping,
        "entities": entities,
        "qa": qa,
    }
    return package, build_preflight(package)


def read_sheet_rows(ws, header_row: int) -> tuple[List[str], List[Dict[str, Any]]]:
    raw_headers = [cell.value for cell in next(ws.iter_rows(min_row=header_row, max_row=header_row))]
    headers = [clean(value) or f"Column {idx + 1}" for idx, value in enumerate(raw_headers)]
    rows: List[Dict[str, Any]] = []
    for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row = {headers[idx]: values[idx] for idx in range(min(len(headers), len(values)))}
        if any(clean(value) for value in row.values()):
            rows.append(row)
    return headers, rows


def validate_selection(
    headers: List[str],
    selected_columns: List[str],
    entity_column: str,
    selected_names: Iterable[str],
    page_size: str,
    qa: List[Dict[str, Any]],
) -> None:
    if not headers:
        qa.append(issue("selection", "error", "sheet가 비어 있거나 헤더를 읽지 못했습니다."))
    if page_size not in {"A4", "B5"}:
        qa.append(issue("selection", "error", "page_size는 A4 또는 B5여야 합니다."))
    if not entity_column:
        qa.append(issue("selection", "error", "기관명 열을 선택하세요."))
    elif entity_column not in headers:
        qa.append(issue("selection", "error", f"기관명 열을 찾지 못했습니다: {entity_column}"))
    if not list(selected_names):
        qa.append(issue("selection", "error", "선택 기관이 없습니다."))
    if not selected_columns:
        qa.append(issue("selection", "error", "사용할 데이터 열을 하나 이상 선택하세요."))
    for column in selected_columns:
        if column not in headers:
            qa.append(issue("selection", "error", f"선택 열을 Excel에서 찾지 못했습니다: {column}"))


def validate_mapping(headers: List[str], selected_columns: List[str], mapping: Dict[str, Any], qa: List[Dict[str, Any]]) -> None:
    selected = set(selected_columns)
    numeric_columns = []
    for slot in mapping.get("kpi_slots", []):
        column = clean(slot.get("value_column"))
        if not column:
            continue
        numeric_columns.append(column)
        validate_mapped_column(column, headers, selected, "KPI", qa)
        compare = clean(slot.get("compare_column"))
        if compare:
            numeric_columns.append(compare)
            validate_mapped_column(compare, headers, selected, "KPI 비교값", qa)
    for slot in mapping.get("chart_slots", []):
        columns = [clean(value) for value in slot.get("value_columns", []) if clean(value)]
        if not columns:
            qa.append(issue("mapping", "error", f"차트 슬롯 값 열이 비어 있습니다: {clean(slot.get('title'))}"))
        for column in columns:
            numeric_columns.append(column)
            validate_mapped_column(column, headers, selected, "차트", qa)
    if len([slot for slot in mapping.get("kpi_slots", []) if clean(slot.get("value_column"))]) < 3:
        qa.append(issue("mapping", "warning", "KPI 슬롯이 3개 미만입니다."))
    if not mapping.get("chart_slots"):
        qa.append(issue("mapping", "warning", "차트 슬롯이 없습니다."))


def validate_mapped_column(column: str, headers: List[str], selected: set[str], label: str, qa: List[Dict[str, Any]]) -> None:
    if column not in headers:
        qa.append(issue("mapping", "error", f"{label} 배정 열을 Excel에서 찾지 못했습니다: {column}"))
    if selected and column not in selected:
        qa.append(issue("mapping", "error", f"{label} 배정 열이 선택 열 목록에 없습니다: {column}"))


def build_entity(entity_id: str, entity_name: str, row: Dict[str, Any], mapping: Dict[str, Any], qa: List[Dict[str, Any]]) -> Dict[str, Any]:
    profile = []
    for item in mapping.get("profile_fields", []):
        label = clean(item.get("label")) or clean(item.get("column"))
        column = clean(item.get("column"))
        if column:
            profile.append({"label": label, "value": clean(row.get(column))})

    kpis = []
    for idx, slot in enumerate(mapping.get("kpi_slots", [])[:6], start=1):
        column = clean(slot.get("value_column"))
        value = row.get(column)
        parsed = to_number(value)
        if column and parsed is None:
            qa.append(issue(entity_name, "error", f"KPI 값이 숫자가 아닙니다: {column}"))
        kpis.append(
            {
                "slot": idx,
                "label": clean(slot.get("label")) or column,
                "value": parsed,
                "display_value": format_value(parsed, clean(slot.get("unit")), int(slot.get("decimals") or 1)),
                "unit": clean(slot.get("unit")),
                "column": column,
            }
        )

    charts = []
    for idx, slot in enumerate(mapping.get("chart_slots", [])[:4], start=1):
        chart_type = clean(slot.get("chart_type")) or "auto"
        columns = [clean(value) for value in slot.get("value_columns", []) if clean(value)]
        labels = [clean(value) for value in slot.get("category_labels", []) if clean(value)]
        points = []
        for point_idx, column in enumerate(columns):
            parsed = to_number(row.get(column))
            if parsed is None:
                qa.append(issue(entity_name, "error", f"차트 값이 숫자가 아닙니다: {column}"))
                continue
            label = labels[point_idx] if point_idx < len(labels) else column
            points.append({"category": label, "value": parsed, "display_value": format_value(parsed, clean(slot.get("unit")), 1)})
        charts.append({"slot": idx, "title": clean(slot.get("title")) or f"차트 {idx}", "chart_type": chart_type, "points": points})

    narrative = clean(mapping.get("narrative_template"))
    if narrative:
        narrative = render_template(narrative, row, entity_name)
    else:
        narrative = default_narrative(entity_name, kpis)
    if not narrative:
        qa.append(issue(entity_name, "warning", "분석문이 비어 있습니다."))

    return {"entity_id": entity_id, "entity_name": entity_name, "profile": profile, "kpis": kpis, "charts": charts, "narrative": narrative}


def build_preflight(package: Dict[str, Any]) -> Dict[str, Any]:
    warnings = [item for item in package["qa"] if item["severity"] == "warning"]
    errors = [item for item in package["qa"] if item["severity"] == "error"]
    return {
        "schema_version": "1.0",
        "status": "blocked" if errors else "ready_with_warnings" if warnings else "ready",
        "summary": {
            "entity_count": len(package.get("entities", [])),
            "qa_warning_count": len(warnings),
            "qa_error_count": len(errors),
        },
        "warnings": warnings,
        "errors": errors,
    }


def default_narrative(entity_name: str, kpis: List[Dict[str, Any]]) -> str:
    visible = [kpi for kpi in kpis if kpi.get("value") is not None][:3]
    if not visible:
        return ""
    values = ", ".join(f"{item['label']} {item['display_value']}" for item in visible)
    return f"{entity_name}의 주요 지표는 {values}로 나타났다."


def render_template(template: str, row: Dict[str, Any], entity_name: str) -> str:
    text = template.replace("{{기관명}}", entity_name)
    for key, value in row.items():
        text = text.replace("{{" + str(key) + "}}", clean(value))
    return text


def format_value(value: float | None, unit: str, decimals: int) -> str:
    if value is None:
        return ""
    if abs(value) >= 1000 and unit not in {"점", "%"}:
        number = f"{value:,.0f}" if decimals == 0 else f"{value:,.{decimals}f}"
    else:
        number = f"{value:.0f}" if decimals == 0 else f"{value:.{decimals}f}"
    return number + unit


def issue(target: str, severity: str, message: str) -> Dict[str, Any]:
    return {"target": target, "severity": severity, "message": message}


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def to_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", "").replace("%", "").strip())
    except ValueError:
        return None


def load_json(path: str | Path) -> Dict[str, Any]:
    return json.loads(Path(path).read_text(encoding="utf-8"))


def run_cli(argv: List[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Build dashboard package and preflight JSON.")
    parser.add_argument("--excel", required=True)
    parser.add_argument("--inspect-output")
    parser.add_argument("--selection")
    parser.add_argument("--mapping")
    parser.add_argument("--page-size", default="A4")
    parser.add_argument("--package-output")
    parser.add_argument("--preflight-output")
    args = parser.parse_args(argv)

    if args.inspect_output:
        output = inspect_workbook(args.excel, args.inspect_output)
        print(str(output))
        return 0

    if not args.selection or not args.mapping or not args.package_output or not args.preflight_output:
        parser.error("--selection, --mapping, --package-output, --preflight-output are required unless --inspect-output is used")

    package, preflight = build_dashboard_package(args.excel, args.selection, args.mapping, args.page_size)
    Path(args.package_output).write_text(json.dumps(package, ensure_ascii=False, indent=2), encoding="utf-8")
    Path(args.preflight_output).write_text(json.dumps(preflight, ensure_ascii=False, indent=2), encoding="utf-8")
    print(str(args.package_output))
    print(str(args.preflight_output))
    return 0


if __name__ == "__main__":
    raise SystemExit(run_cli())
