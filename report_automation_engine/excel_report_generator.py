"""Excel 집계표 기반 보고서 본문 생성 엔진.

이 모듈은 사용자가 추가로 제공한 `excel_report_generator_with_style.py`를
프로젝트 내부 보조 엔진으로 편입한 것이다. 기존 VBA add-in은 엑셀 파일 안에
보고서용 산출 시트를 만드는 역할을 맡고, 이 Python 엔진은 원본 집계표 또는
산출 시트를 읽어 사람이 검토할 수 있는 본문 초안을 TXT로 생성하는 역할을 맡긴다.

설계상 중요한 점:
- 스타일과 문장 패턴은 JSON 설정(`config/default_style_schema.json`)으로 분리한다.
- 표 추출, 핵심 수치 랭킹, 배너별 비교 문장을 한 파일에서 처리하지만, 각 단계는
  함수로 나뉘어 있어 이후 C# 런처나 VBA 산출 시트와 연결하기 쉽다.
- 이 코드는 아직 "완성 보고서"가 아니라 "검토 가능한 초안" 생성을 목표로 한다.
"""

from __future__ import annotations

import json
import re
import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox


# =========================
# Config loading
# =========================
REQUIRED_TOP_LEVEL_KEYS = {
    "schema_version",
    "report_style",
    "title_rules",
    "ranking_rules",
    "table_rules",
    "characteristic_rules",
}


@dataclass
class StyleConfig:
    """보고서 문체/표 해석 규칙을 담는 설정 객체.

    JSON을 그대로 dict로 들고 있는 이유는 규칙 항목이 계속 늘어날 가능성이 높기
    때문이다. 지금 단계에서는 엄격한 세부 dataclass보다 확장성이 더 중요하다.
    """

    schema_version: str
    report_style: Dict[str, Any]
    title_rules: Dict[str, Any]
    ranking_rules: Dict[str, Any]
    table_rules: Dict[str, Any]
    characteristic_rules: Dict[str, Any]


@dataclass
class ExcelTableRow:
    """집계표의 데이터 행 하나를 표준화한 구조.

    label/subgroup/group을 분리해 두면 전체 행, 지역별 행, 성별/연령별 행을 같은
    문장 생성 로직에서 다룰 수 있다. values에는 응답 항목명 -> 수치가 들어간다.
    """

    label: str
    group: str
    subgroup: str
    case_count: Optional[float]
    values: Dict[str, Optional[float]]


@dataclass
class ExcelTableData:
    """하나의 집계표 블록을 보고서 생성에 필요한 형태로 정규화한 구조."""

    sheet_name: str
    title: str
    normalized_title: str
    subtype: str
    start_row: int
    end_row: int
    unit_text: str
    response_columns: List[str]
    ranked_columns: List[str]
    summary_columns: List[str]
    mean_columns: List[str]
    score_columns: List[str]
    rows: List[ExcelTableRow]
    total_row: Optional[ExcelTableRow]
    multi_response: bool


# =========================
# Default runtime constants
# =========================
MAX_TITLE_LOOKBACK = 5
MAX_EMPTY_STREAK = 4
DEFAULT_OUTPUT_SUFFIX = "_자동생성.txt"
TITLE_UNKNOWN = "제목 미확인"
DEFAULT_CONFIG_PATH = Path(__file__).resolve().parent / "config" / "default_style_schema.json"


# =========================
# Validation helpers
# =========================
def _require_keys(payload: Dict[str, Any], keys: set[str], scope: str) -> None:
    missing = sorted(key for key in keys if key not in payload)
    if missing:
        raise ValueError(f"{scope} 필수 키 누락: {', '.join(missing)}")


def _require_list(payload: Dict[str, Any], key: str, scope: str) -> List[str]:
    value = payload.get(key)
    if not isinstance(value, list):
        raise ValueError(f"{scope}.{key} 는 리스트여야 합니다.")
    return [str(item) for item in value]


def load_style_config(path: str | Path) -> StyleConfig:
    config_path = Path(path)
    payload = json.loads(config_path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError("스타일 설정 파일은 JSON 객체여야 합니다.")

    _require_keys(payload, REQUIRED_TOP_LEVEL_KEYS, "root")

    for key in REQUIRED_TOP_LEVEL_KEYS - {"schema_version"}:
        if not isinstance(payload[key], dict):
            raise ValueError(f"root.{key} 는 객체여야 합니다.")

    report_style = payload["report_style"]
    title_rules = payload["title_rules"]
    ranking_rules = payload["ranking_rules"]
    table_rules = payload["table_rules"]
    characteristic_rules = payload["characteristic_rules"]

    _require_keys(
        report_style,
        {
            "main_bullet",
            "sub_bullet",
            "top_pattern",
            "follow_pattern",
            "mean_pattern",
            "score_pattern",
            "characteristic_pattern",
            "district_pattern",
        },
        "report_style",
    )
    _require_keys(
        title_rules,
        {"strip_regexes", "strip_prefixes", "strip_markers", "skip_title_keywords", "skip_title_exact"},
        "title_rules",
    )
    _require_keys(
        ranking_rules,
        {
            "exclude_labels",
            "exclude_column_keywords",
            "max_follow_items",
            "ignore_zero",
            "deduplicate_labels",
            "multi_response_threshold",
        },
        "ranking_rules",
    )
    _require_keys(
        table_rules,
        {
            "case_column_name",
            "total_label",
            "mean_keywords",
            "summary_keywords",
            "score_keywords",
            "unit_keywords",
            "percent_keywords",
            "base_keywords",
        },
        "table_rules",
    )
    _require_keys(characteristic_rules, {"preferred_groups", "max_groups"}, "characteristic_rules")

    _require_list(title_rules, "strip_regexes", "title_rules")
    _require_list(title_rules, "strip_prefixes", "title_rules")
    _require_list(title_rules, "strip_markers", "title_rules")
    _require_list(title_rules, "skip_title_keywords", "title_rules")
    _require_list(title_rules, "skip_title_exact", "title_rules")
    _require_list(ranking_rules, "exclude_labels", "ranking_rules")
    _require_list(ranking_rules, "exclude_column_keywords", "ranking_rules")
    _require_list(table_rules, "mean_keywords", "table_rules")
    _require_list(table_rules, "summary_keywords", "table_rules")
    _require_list(table_rules, "score_keywords", "table_rules")
    _require_list(table_rules, "unit_keywords", "table_rules")
    _require_list(table_rules, "percent_keywords", "table_rules")
    _require_list(table_rules, "base_keywords", "table_rules")
    _require_list(characteristic_rules, "preferred_groups", "characteristic_rules")

    return StyleConfig(
        schema_version=str(payload["schema_version"]),
        report_style=report_style,
        title_rules=title_rules,
        ranking_rules=ranking_rules,
        table_rules=table_rules,
        characteristic_rules=characteristic_rules,
    )


# =========================
# Common utils
# =========================
def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def parse_number(value: Any) -> Optional[float]:
    text = clean_text(value)
    if not text or text == "-":
        return None

    text = text.replace(",", "").replace("%", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group())
    except ValueError:
        return None


def format_number(value: Optional[float], digits: int = 1) -> str:
    if value is None:
        return "-"
    if abs(value - round(value)) < 1e-9:
        return f"{int(round(value)):,}"
    return f"{value:,.{digits}f}"


def format_percent(value: Optional[float]) -> str:
    if value is None:
        return "-"
    return f"{format_number(value, 1)}%"


def render_pattern(pattern: str, **kwargs: Any) -> str:
    return pattern.format(**kwargs)


def fill_merged_cells(ws) -> None:
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        min_col = merged_range.min_col
        min_row = merged_range.min_row
        max_col = merged_range.max_col
        max_row = merged_range.max_row

        top_left_value = ws.cell(min_row, min_col).value
        if top_left_value is None:
            continue

        ws.unmerge_cells(str(merged_range))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if ws.cell(r, c).value is None:
                    ws.cell(r, c).value = top_left_value


def row_values(ws, row_idx: int) -> List[str]:
    return [clean_text(ws.cell(row=row_idx, column=c).value) for c in range(1, ws.max_column + 1)]


def join_row_text(values: List[str]) -> str:
    return " ".join(v for v in values if v).strip()


def unique_preserve_order(items: List[str]) -> List[str]:
    seen = set()
    result = []
    for item in items:
        if item and item not in seen:
            seen.add(item)
            result.append(item)
    return result


def is_effectively_empty_row(values: List[str]) -> bool:
    return all(not v for v in values)


# =========================
# Config-aware label/title helpers
# =========================
def is_total_label(text: str, config: StyleConfig) -> bool:
    return clean_text(text).replace(" ", "") == clean_text(config.table_rules["total_label"]).replace(" ", "")


def is_base_row(values: List[str], config: StyleConfig) -> bool:
    keywords = [clean_text(x) for x in config.table_rules["base_keywords"]]
    for v in values:
        if not v:
            continue
        if any(v.startswith(k) for k in keywords if k):
            return True
    return False


def is_case_header_row(values: List[str], config: StyleConfig) -> bool:
    target = clean_text(config.table_rules["case_column_name"])
    return target in join_row_text(values)


def is_unit_row(values: List[str], config: StyleConfig) -> bool:
    joined = join_row_text(values)
    unit_keywords = [clean_text(x) for x in config.table_rules["unit_keywords"]]
    percent_keywords = [clean_text(x) for x in config.table_rules["percent_keywords"]]
    if any(k and k in joined for k in unit_keywords):
        return True
    if joined in percent_keywords:
        return True
    return False


def normalize_label(text: str) -> str:
    text = clean_text(text)
    text = text.replace("■ ", "").replace("■", "").strip()
    return text


def should_exclude_label(label: str, config: StyleConfig) -> bool:
    normalized = normalize_label(label)
    excludes = {normalize_label(x) for x in config.ranking_rules["exclude_labels"]}
    return normalized in excludes


def should_exclude_column(col_name: str, config: StyleConfig) -> bool:
    normalized = clean_text(col_name)
    if not normalized:
        return True

    exclude_keywords = [clean_text(x) for x in config.ranking_rules["exclude_column_keywords"]]
    return any(keyword and keyword in normalized for keyword in exclude_keywords)


def dedupe_title_tokens(text: str) -> str:
    text = clean_text(text)
    if not text:
        return text

    parts = re.split(r"\s*(?=(?:\[표|<표))", text)
    parts = [clean_text(p) for p in parts if clean_text(p)]
    parts = unique_preserve_order(parts)
    if parts:
        return " ".join(parts)

    words = text.split()
    deduped: List[str] = []
    for word in words:
        if not deduped or deduped[-1] != word:
            deduped.append(word)
    return " ".join(deduped)


def skip_candidate_title(text: str, config: StyleConfig) -> bool:
    text = clean_text(text)
    if not text:
        return True

    if text in {clean_text(x) for x in config.title_rules["skip_title_exact"]}:
        return True

    for keyword in config.title_rules["skip_title_keywords"]:
        keyword = clean_text(keyword)
        if keyword and keyword in text:
            return True

    if is_case_header_row([text], config):
        return True
    if is_base_row([text], config):
        return True
    if text.endswith("나타남"):
        return True
    return False


def strip_title_prefix(title: str, config: StyleConfig) -> str:
    title = dedupe_title_tokens(title)
    title = clean_text(title)

    for regex in config.title_rules["strip_regexes"]:
        title = re.sub(regex, "", title)

    for prefix in config.title_rules["strip_prefixes"]:
        prefix = clean_text(prefix)
        if prefix and title.startswith(prefix):
            title = title[len(prefix) :]

    for marker in config.title_rules["strip_markers"]:
        marker = clean_text(marker)
        pos = title.find(marker)
        if marker and pos > 0:
            title = title[:pos]
            break

    return clean_text(title)


def normalize_analysis_title(title: str, config: StyleConfig) -> str:
    title = strip_title_prefix(title, config)
    title = re.sub(r"\s+사업체 특성별$", "", title)
    title = re.sub(r"\s+사업체 특성$", "", title)
    return clean_text(title)


def find_title_for_table(ws, header_row_idx: int, config: StyleConfig) -> str:
    candidates: List[str] = []
    for r in range(max(1, header_row_idx - MAX_TITLE_LOOKBACK), header_row_idx):
        values = row_values(ws, r)
        joined = dedupe_title_tokens(join_row_text(values))
        if not joined:
            continue

        if "[표" in joined or "<표" in joined:
            return joined

        first_nonempty = next((v for v in values if v), "")
        first_nonempty = dedupe_title_tokens(first_nonempty)
        if first_nonempty and not skip_candidate_title(first_nonempty, config):
            candidates.append(first_nonempty)

        if joined and not skip_candidate_title(joined, config):
            candidates.append(joined)

    return candidates[-1] if candidates else TITLE_UNKNOWN


# =========================
# Table boundary / header detection
# =========================
def find_table_starts(ws, config: StyleConfig) -> List[int]:
    starts = []
    for r in range(1, ws.max_row + 1):
        values = row_values(ws, r)
        if is_case_header_row(values, config):
            starts.append(r)
    return starts


def find_table_starts_fast(ws, config: StyleConfig) -> List[int]:
    """병합셀 보정 전에 표 시작 행 후보를 빠르게 찾는다.

    원본 제공 코드는 모든 시트에 대해 `fill_merged_cells()`를 먼저 실행했다.
    이 방식은 정확도는 좋지만, 목차/빈 시트/매우 큰 시트가 섞인 파일에서는
    불필요한 셀 쓰기와 unmerge 작업이 많아진다.

    여기서는 먼저 읽기 전용에 가까운 방식으로 `사례수` 헤더가 있는 행만 찾고,
    후보가 있는 시트에 대해서만 병합셀 보정을 수행한다. 병합셀 보정 이후에는
    다시 `find_table_starts()`를 호출해 정확한 시작 행을 확정한다.
    """

    target = clean_text(config.table_rules["case_column_name"])
    starts: List[int] = []
    if not target:
        return starts

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # values_only=True는 Cell 객체를 만들지 않아 대형 집계표에서 훨씬 가볍다.
        values = [clean_text(value) for value in row]
        if target in join_row_text(values):
            starts.append(row_idx)
    return starts


def find_table_end(ws, start_row: int, config: StyleConfig) -> int:
    empty_streak = 0
    for r in range(start_row + 1, ws.max_row + 1):
        values = row_values(ws, r)
        joined = join_row_text(values)

        if joined:
            if "[표" in joined or "<표" in joined:
                return r - 1
            if is_case_header_row(values, config):
                return r - 1

        if is_effectively_empty_row(values):
            empty_streak += 1
            if empty_streak >= MAX_EMPTY_STREAK:
                return r - empty_streak
        else:
            empty_streak = 0
    return ws.max_row


def find_case_col_idx(header_values: List[str], config: StyleConfig) -> Optional[int]:
    target = clean_text(config.table_rules["case_column_name"])
    for i, v in enumerate(header_values):
        if target in v:
            return i
    return None


def find_upper_header_row(ws, header_row_idx: int, config: StyleConfig) -> Optional[int]:
    if header_row_idx <= 1:
        return None
    upper = row_values(ws, header_row_idx - 1)
    if is_effectively_empty_row(upper):
        return None
    if is_base_row(upper, config):
        return None

    joined = join_row_text(upper)
    if is_unit_row(upper, config):
        return header_row_idx - 1

    nonempty = sum(1 for x in upper if x)
    if 1 <= nonempty <= 12 and joined:
        return header_row_idx - 1
    return None


def classify_column_name(col_name: str, config: StyleConfig) -> str:
    name = clean_text(col_name)
    if not name:
        return "ignore"

    if should_exclude_column(name, config):
        return "ignore"

    mean_keywords = [clean_text(x) for x in config.table_rules["mean_keywords"]]
    summary_keywords = [clean_text(x) for x in config.table_rules["summary_keywords"]]
    score_keywords = [clean_text(x) for x in config.table_rules["score_keywords"]]
    percent_keywords = [clean_text(x) for x in config.table_rules["percent_keywords"]]

    if any(k and k in name for k in mean_keywords):
        return "mean"
    if any(k and k in name for k in score_keywords):
        return "score"
    if any(k and k in name for k in summary_keywords):
        return "summary"
    if name in percent_keywords:
        return "ignore"
    return "rank"


def build_response_columns(ws, header_row_idx: int, case_col_idx: int, config: StyleConfig) -> Tuple[List[str], str]:
    header_values = row_values(ws, header_row_idx)
    upper_header_idx = find_upper_header_row(ws, header_row_idx, config)
    upper_values = row_values(ws, upper_header_idx) if upper_header_idx else []

    unit_text = ""
    if upper_values and is_unit_row(upper_values, config):
        unit_text = join_row_text(upper_values)

    response_columns: List[str] = []
    for c in range(case_col_idx + 1, len(header_values)):
        lower = clean_text(header_values[c])
        upper = clean_text(upper_values[c]) if upper_values and c < len(upper_values) else ""

        if not lower and not upper:
            continue

        if lower == upper and lower:
            col_name = lower
        elif upper and lower and upper != lower and upper not in config.table_rules["percent_keywords"]:
            col_name = f"{upper}_{lower}"
        else:
            col_name = lower or upper

        col_name = clean_text(col_name)
        if col_name:
            response_columns.append(col_name)

    response_columns = unique_preserve_order(response_columns)
    return response_columns, unit_text


# =========================
# Row parsing
# =========================
def split_row_label_parts(left_cells: List[str]) -> Tuple[str, str, str]:
    cleaned = [normalize_label(cell) for cell in left_cells if normalize_label(cell)]
    if not cleaned:
        return "", "", ""
    if len(cleaned) == 1:
        return cleaned[0], "", cleaned[0]
    group = cleaned[0]
    subgroup = cleaned[-1]
    if group == subgroup:
        subgroup = ""
    label = subgroup or group
    return label, group, subgroup


def detect_table_subtype(title: str, rows: List[ExcelTableRow], config: StyleConfig) -> str:
    raw = strip_title_prefix(title, config)
    groups = {row.group for row in rows if row.group}

    district_hints = ["진흥지구", "지역별"]
    if any(h in raw for h in district_hints):
        return "district"
    if any("특성" in raw for h in [raw]):
        return "characteristic"
    if any("진흥지구" in g for g in groups):
        return "district"
    if any(any(keyword in g for keyword in config.characteristic_rules["preferred_groups"]) for g in groups):
        return "characteristic"
    return "overall"


def build_table_data_from_sheet(ws, header_row_idx: int, config: StyleConfig) -> Optional[ExcelTableData]:
    header_values = row_values(ws, header_row_idx)
    case_col_idx = find_case_col_idx(header_values, config)
    if case_col_idx is None:
        return None

    title = find_title_for_table(ws, header_row_idx, config)
    response_columns, unit_text = build_response_columns(ws, header_row_idx, case_col_idx, config)
    if not response_columns:
        return None

    ranked_columns = [c for c in response_columns if classify_column_name(c, config) == "rank"]
    summary_columns = [c for c in response_columns if classify_column_name(c, config) == "summary"]
    mean_columns = [c for c in response_columns if classify_column_name(c, config) == "mean"]
    score_columns = [c for c in response_columns if classify_column_name(c, config) == "score"]

    start_row = header_row_idx
    end_row = find_table_end(ws, header_row_idx, config)

    rows: List[ExcelTableRow] = []
    total_row: Optional[ExcelTableRow] = None

    for r in range(header_row_idx + 1, end_row + 1):
        values = row_values(ws, r)
        joined = join_row_text(values)

        if is_effectively_empty_row(values):
            continue
        if is_base_row(values, config):
            continue
        if is_unit_row(values, config):
            continue
        if is_case_header_row(values, config):
            continue
        if "[표" in joined or "<표" in joined:
            break

        left_cells = values[:case_col_idx]
        label, group, subgroup = split_row_label_parts(left_cells)
        case_count = parse_number(values[case_col_idx]) if case_col_idx < len(values) else None

        value_map: Dict[str, Optional[float]] = {}
        source_values = values[case_col_idx + 1 : case_col_idx + 1 + len(response_columns)]
        for col_name, raw_value in zip(response_columns, source_values):
            value_map[col_name] = parse_number(raw_value)

        if not label and case_count is None and all(v is None for v in value_map.values()):
            continue

        row = ExcelTableRow(
            label=label,
            group=group,
            subgroup=subgroup,
            case_count=case_count,
            values=value_map,
        )
        rows.append(row)
        if is_total_label(label, config):
            total_row = row

    if not rows:
        return None

    multi_threshold = float(config.ranking_rules["multi_response_threshold"])
    numeric_total = 0.0
    if total_row is not None:
        numeric_total = sum((total_row.values.get(c) or 0.0) for c in ranked_columns)

    subtype = detect_table_subtype(title, rows, config)
    return ExcelTableData(
        sheet_name=ws.title,
        title=strip_title_prefix(title, config),
        normalized_title=normalize_analysis_title(title, config),
        subtype=subtype,
        start_row=start_row,
        end_row=end_row,
        unit_text=unit_text,
        response_columns=response_columns,
        ranked_columns=ranked_columns,
        summary_columns=summary_columns,
        mean_columns=mean_columns,
        score_columns=score_columns,
        rows=rows,
        total_row=total_row,
        multi_response=(numeric_total > multi_threshold) if ranked_columns else False,
    )


def extract_tables_from_excel(
    file_path: str | Path,
    config: StyleConfig,
    sheet_name: Optional[str] = None,
    max_tables: Optional[int] = None,
) -> List[ExcelTableData]:
    """엑셀 파일에서 보고서 문장 생성에 사용할 표 블록을 추출한다.

    `sheet_name`과 `max_tables`는 런처 연동을 위해 추가한 안전장치다.
    대형 집계표를 알파 단계에서 전부 읽으면 사용자가 "멈췄다"고 느낄 수 있으므로,
    런처는 우선 일부 표만 빠르게 초안화하고 후속 알파에서 성능을 더 다듬는다.
    """

    wb = openpyxl.load_workbook(file_path, data_only=True)
    tables: List[ExcelTableData] = []
    for ws in wb.worksheets:
        if sheet_name and ws.title != sheet_name:
            continue
        if ws.title.strip().upper() == "INDEX":
            continue
        if ws.max_row == 0 or ws.max_column == 0:
            continue
        fast_starts = find_table_starts_fast(ws, config)
        if not fast_starts:
            continue
        fill_merged_cells(ws)
        starts = find_table_starts(ws, config)
        for start_row in starts:
            table = build_table_data_from_sheet(ws, start_row, config)
            if table is not None:
                tables.append(table)
                if max_tables is not None and len(tables) >= max_tables:
                    return tables
    return tables


# =========================
# Ranking / rendering
# =========================
def choose_rank_columns(table: ExcelTableData, config: StyleConfig) -> List[str]:
    """보고서 순위 문장에 사용할 응답 열만 고른다.

    원본 제공 코드에는 이 함수가 `GLOBAL_CONFIG`를 직접 참조했다. GUI 실행만
    고려하면 문제가 드러나지 않지만, 런처나 테스트 코드에서 이 모듈을 import해
    사용할 때는 전역 값이 세팅되지 않아 실패할 수 있다. 그래서 프로젝트 편입 시
    설정 객체를 명시적으로 받도록 바꿨다.
    """

    return [c for c in table.ranked_columns if not should_exclude_column(c, config=config)]


def rank_columns(row: ExcelTableRow, columns: List[str], config: StyleConfig) -> List[Tuple[str, float]]:
    ranked: List[Tuple[str, float]] = []
    ignore_zero = bool(config.ranking_rules["ignore_zero"])

    for col in columns:
        if should_exclude_column(col, config):
            continue
        value = row.values.get(col)
        if value is None:
            continue
        if ignore_zero and abs(value) < 1e-12:
            continue
        ranked.append((col, value))

    if bool(config.ranking_rules["deduplicate_labels"]):
        dedup: Dict[str, float] = {}
        for name, value in ranked:
            dedup.setdefault(name, value)
        ranked = list(dedup.items())

    ranked.sort(key=lambda x: (-x[1], x[0]))
    return ranked


def format_items(items: List[Tuple[str, float]], use_percent: bool = True) -> str:
    formatted = []
    for name, value in items:
        value_text = format_percent(value) if use_percent else format_number(value, 1)
        formatted.append(f"'{name}'({value_text})")
    return ", ".join(formatted)


def render_mean_line(table: ExcelTableData, config: StyleConfig) -> Optional[str]:
    if table.total_row is None or not table.mean_columns:
        return None
    first_mean = table.mean_columns[0]
    return render_pattern(
        config.report_style["mean_pattern"],
        title=first_mean,
        mean_value=format_number(table.total_row.values.get(first_mean), 1),
    )


def render_score_line(table: ExcelTableData, config: StyleConfig) -> Optional[str]:
    if table.total_row is None or not table.score_columns:
        return None

    score_5_value = None
    score_100_value = None
    for col in table.score_columns:
        if "5점" in col:
            score_5_value = format_number(table.total_row.values.get(col), 1)
        elif "100점" in col:
            score_100_value = format_number(table.total_row.values.get(col), 1)

    if score_5_value is None and score_100_value is None:
        return None

    return render_pattern(
        config.report_style["score_pattern"],
        title=table.normalized_title,
        score_5_value=score_5_value or "-",
        score_100_value=score_100_value or "-",
    )


def overall_summary(table: ExcelTableData, config: StyleConfig) -> str:
    if table.total_row is None:
        return f"{table.normalized_title} 문항의 전체 응답을 요약하지 못했습니다."

    ranked = rank_columns(table.total_row, table.ranked_columns, config)
    if ranked:
        top1 = ranked[0]
        follow_n = int(config.ranking_rules["max_follow_items"])
        others = ranked[1 : 1 + follow_n]
        line1 = render_pattern(
            config.report_style["top_pattern"],
            title=table.normalized_title,
            top_label=top1[0],
            top_value=format_number(top1[1], 1),
        )
        lines = [line1]
        if others:
            lines.append(
                render_pattern(
                    config.report_style["follow_pattern"],
                    items=format_items(others, use_percent=True),
                )
            )
    else:
        lines = [f"{table.normalized_title} 문항의 비율형 응답을 확인하지 못했습니다."]

    score_line = render_score_line(table, config)
    if score_line:
        lines.append(score_line)

    mean_line = render_mean_line(table, config)
    if mean_line:
        lines.append(mean_line)

    if table.summary_columns and table.total_row is not None:
        summary_parts = []
        for col in table.summary_columns[:2]:
            val = table.total_row.values.get(col)
            if val is not None:
                summary_parts.append(f"{col} {format_number(val, 1)}")
        if summary_parts:
            lines.append("참고로 " + ", ".join(summary_parts) + "으로 집계됨")

    return "\n".join(lines)


def district_summary(table: ExcelTableData, config: StyleConfig) -> str:
    parts = []
    for row in table.rows:
        if is_total_label(row.label, config):
            continue
        if should_exclude_label(row.label, config):
            continue
        ranked = rank_columns(row, table.ranked_columns, config)
        if ranked:
            parts.append(f"{row.label} '{ranked[0][0]}'({format_percent(ranked[0][1])})")

    if not parts:
        return ""
    return render_pattern(config.report_style["district_pattern"], items=", ".join(parts[:6]))


def select_characteristic_groups(table: ExcelTableData, config: StyleConfig) -> List[str]:
    preferred = [clean_text(x) for x in config.characteristic_rules["preferred_groups"]]
    max_groups = int(config.characteristic_rules["max_groups"])

    groups: List[str] = []
    seen = set()

    for wanted in preferred:
        for row in table.rows:
            if row.group == wanted and wanted not in seen:
                groups.append(wanted)
                seen.add(wanted)

    if len(groups) >= max_groups:
        return groups[:max_groups]

    for row in table.rows:
        if row.group and row.group not in seen:
            groups.append(row.group)
            seen.add(row.group)
        if len(groups) >= max_groups:
            break
    return groups[:max_groups]


def characteristic_summary(table: ExcelTableData, config: StyleConfig) -> str:
    if table.total_row is None:
        return ""

    ranked_total = rank_columns(table.total_row, table.ranked_columns, config)
    if not ranked_total:
        return ""

    focus_label = ranked_total[0][0]
    descriptions = []
    for group in select_characteristic_groups(table, config):
        candidates = [
            row for row in table.rows
            if row.group == group and not is_total_label(row.label, config) and not should_exclude_label(row.label, config)
        ]
        if not candidates:
            continue
        best = max(candidates, key=lambda x: x.values.get(focus_label) or -1)
        subgroup = best.subgroup or best.label
        descriptions.append(f"{group} '{subgroup}'({format_percent(best.values.get(focus_label))})")

    if not descriptions:
        return ""

    return render_pattern(
        config.report_style["characteristic_pattern"],
        focus_label=focus_label,
        items=", ".join(descriptions),
    )


def group_tables_by_title(tables: List[ExcelTableData]) -> List[Tuple[str, List[ExcelTableData]]]:
    groups: List[Tuple[str, List[ExcelTableData]]] = []
    current_title = ""
    current_tables: List[ExcelTableData] = []

    sorted_tables = sorted(tables, key=lambda x: (x.sheet_name, x.start_row))
    for table in sorted_tables:
        if current_title and table.normalized_title != current_title:
            groups.append((current_title, current_tables))
            current_tables = []
        current_title = table.normalized_title
        current_tables.append(table)

    if current_tables:
        groups.append((current_title, current_tables))
    return groups


def render_question_section(title: str, tables: List[ExcelTableData], config: StyleConfig) -> str:
    overall_table = next((t for t in tables if t.subtype == "overall"), None)
    district_table = next((t for t in tables if t.subtype == "district"), None)
    characteristic_table = next((t for t in tables if t.subtype == "characteristic"), None)

    primary = overall_table or district_table or characteristic_table
    if primary is None:
        return ""

    main_bullet = config.report_style["main_bullet"]
    lines = [f"▶ {title}", "", f"{main_bullet} {overall_summary(primary, config)}"]

    if district_table is not None:
        text = district_summary(district_table, config)
        if text:
            lines.extend(["", f"{main_bullet} {text}"])

    if characteristic_table is not None:
        text = characteristic_summary(characteristic_table, config)
        if text:
            lines.extend(["", f"{main_bullet} {text}"])

    return "\n".join(lines).strip()


def render_excel_report(
    file_path: str | Path,
    config: StyleConfig,
    sheet_name: Optional[str] = None,
    max_tables: Optional[int] = None,
) -> str:
    tables = extract_tables_from_excel(file_path, config, sheet_name=sheet_name, max_tables=max_tables)
    sections = []
    for title, grouped_tables in group_tables_by_title(tables):
        section = render_question_section(title, grouped_tables, config)
        if section:
            sections.append(section)
    file_name = Path(file_path).name
    return f"{file_name} 자동 생성 보고서 본문\n\n" + "\n\n".join(sections)


def render_from_generated_output_sheets(file_path: str | Path) -> Optional[str]:
    """VBA add-in이 만든 `보고서_분석문` 시트에서 TXT 초안을 만든다.

    런처 기반 알파의 기본 경로는 `Excel 산출 시트 생성 → 그 산출 시트를 사람이 검토`
    이다. 따라서 Python이 대형 원본 집계표를 처음부터 다시 해석하는 것보다, 이미
    VBA가 정리한 `보고서_분석문` 시트를 읽는 편이 훨씬 빠르고 사용자 기대에도 맞다.

    반환값이 None이면 아직 산출 시트가 없다는 뜻이므로, 기존 원본 파서로 fallback한다.
    """

    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    narrative_sheets = [ws for ws in wb.worksheets if ws.title.startswith("보고서_분석문")]
    if not narrative_sheets:
        return None

    ws = narrative_sheets[-1]
    lines: List[str] = [f"{Path(file_path).name} 자동 생성 보고서 본문", ""]
    for row in ws.iter_rows(min_row=2, values_only=True):
        table_key = clean_text(row[0] if len(row) > 0 else "")
        title = clean_text(row[1] if len(row) > 1 else "")
        narrative = clean_text(row[8] if len(row) > 8 and row[8] else row[2] if len(row) > 2 else "")
        if not title and not narrative:
            continue
        if title:
            lines.append(f"▶ {title}")
        if narrative:
            # 셀 안의 줄바꿈은 유지하되, 각 표 사이에는 빈 줄을 넣어 검토하기 쉽게 한다.
            lines.append(narrative.replace("\r\n", "\n").replace("\r", "\n"))
        if table_key:
            lines.append(f"[source: {table_key}]")
        lines.append("")

    return "\n".join(lines).strip() + "\n"


# =========================
# GUI app
# =========================
GLOBAL_CONFIG: StyleConfig


def run_cli(argv: Optional[List[str]] = None) -> int:
    """런처/배치 실행용 CLI 진입점.

    기존 제공 코드는 tkinter 파일 선택 창을 띄우는 방식이었다. 런처 기반 알파에서는
    C# 프로그램이 이 스크립트를 별도 프로세스로 호출해야 하므로, GUI 없이 입력/출력
    경로를 인자로 받는 CLI가 필요하다.
    """

    parser = argparse.ArgumentParser(description="엑셀 집계표에서 보고서 본문 초안을 생성합니다.")
    parser.add_argument("--excel", required=True, help="입력 엑셀 집계표 경로")
    parser.add_argument("--config", default=str(DEFAULT_CONFIG_PATH), help="스타일 설정 JSON 경로")
    parser.add_argument("--output", help="출력 TXT 경로. 생략 시 엑셀 파일 옆에 _draft.txt로 저장")
    parser.add_argument("--sheet", help="분석할 시트명. 생략 시 모든 후보 시트")
    parser.add_argument("--max-tables", type=int, default=30, help="알파 미리보기용 최대 표 개수")
    parser.add_argument("--raw-tables", action="store_true", help="보고서_분석문 산출 시트 대신 원본 표를 직접 파싱")
    args = parser.parse_args(argv)

    excel_path = Path(args.excel)
    output_path = Path(args.output) if args.output else excel_path.with_name(excel_path.stem + "_draft.txt")
    report = None if args.raw_tables else render_from_generated_output_sheets(excel_path)
    if report is None:
        config = load_style_config(args.config)
        report = render_excel_report(excel_path, config, sheet_name=args.sheet, max_tables=args.max_tables)
    output_path.write_text(report, encoding="utf-8")
    print(str(output_path))
    return 0


def ask_for_config_file() -> Optional[str]:
    return filedialog.askopenfilename(
        title="스타일 설정 JSON 선택",
        filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
    )


def ask_for_excel_files() -> Tuple[str, ...]:
    return filedialog.askopenfilenames(
        title="분석할 엑셀 파일 선택(복수 선택 가능)",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
    )


def main() -> None:
    global GLOBAL_CONFIG

    root = tk.Tk()
    root.withdraw()

    config_path = ask_for_config_file()
    if not config_path:
        # 스타일 JSON을 매번 선택하는 방식은 테스트 단계에서는 번거롭다.
        # 사용자가 선택 창을 닫으면 프로젝트에 포함된 기본 스타일을 사용한다.
        # 단, 기본 파일이 없으면 설정 없이 문장을 생성할 수 없으므로 종료한다.
        if DEFAULT_CONFIG_PATH.exists():
            config_path = str(DEFAULT_CONFIG_PATH)
        else:
            messagebox.showinfo("취소", "스타일 설정 파일 선택이 취소되었습니다.")
            return

    try:
        GLOBAL_CONFIG = load_style_config(config_path)
    except Exception as e:
        messagebox.showerror("설정 오류", f"스타일 설정 로드 실패\n\n{e}")
        return

    file_paths = ask_for_excel_files()
    if not file_paths:
        messagebox.showinfo("취소", "엑셀 파일 선택이 취소되었습니다.")
        return

    success_files = []
    failed_files = []
    for file_path in file_paths:
        try:
            report = render_excel_report(file_path, GLOBAL_CONFIG)
            output_path = Path(file_path).with_name(Path(file_path).stem + DEFAULT_OUTPUT_SUFFIX)
            output_path.write_text(report, encoding="utf-8")
            success_files.append(str(output_path))
        except Exception as e:
            failed_files.append(f"{Path(file_path).name}: {e}")

    msg_parts = []
    if success_files:
        msg_parts.append("완료된 파일\n\n" + "\n".join(success_files[:20]))
    if failed_files:
        msg_parts.append("실패한 파일\n\n" + "\n".join(failed_files[:20]))

    messagebox.showinfo("처리 결과", "\n\n".join(msg_parts) if msg_parts else "처리 결과가 없습니다.")


if __name__ == "__main__":
    import sys

    # 인자가 있으면 런처/배치용 CLI로 동작하고, 인자가 없으면 기존처럼 간단 GUI로 동작한다.
    if len(sys.argv) > 1:
        raise SystemExit(run_cli())
    main()
